import openpyxl
from fuzzywuzzy import fuzz, process
import sys

# Ruta del archivo Excel
excel_file = 'src/app/backend/Demografia.xlsx'

# Función para combinar los encabezados
def combinar_encabezados(sheet):
    encabezados = []
    encabezado_actual = ""

    for col in range(1, sheet.max_column + 1):
        valor_fila_1 = sheet.cell(row=1, column=col).value
        valor_fila_2 = sheet.cell(row=2, column=col).value

        if valor_fila_1:
            encabezado_actual = valor_fila_1.lower()
        
        if valor_fila_2:
            encabezados.append(f"{encabezado_actual} ({valor_fila_2.lower()})")
        else:
            encabezados.append(encabezado_actual)

    return [encabezado.strip() for encabezado in encabezados]

# Función para obtener la respuesta basada en la pregunta
def obtener_respuesta(pregunta):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active  # Seleccionar la hoja activa (por defecto, la primera)

    # Combinar encabezados
    encabezados = combinar_encabezados(sheet)

    # Normalizar encabezados
    encabezados = [encabezado.strip().lower() for encabezado in encabezados]

    # Limpiar y normalizar la pregunta
    pregunta = pregunta.lower().strip()

    # Inicializar variables para almacenar la información relevante
    respuesta = None
    categoria = None

    # Mapear las categorías con los encabezados exactos del archivo
    categorias_mapeadas = {
        "esperanza de vida hombre": "esperanza de vida al nacer (años) (hombre)",
        "esperanza de vida mujer": "esperanza de vida al nacer (años) (mujer)",
        "composición 0 a 14": "composición de la población en años (porcentaje) (0 a 14)",
        "composición 15 a 64": "composición de la población en años (porcentaje) (15 a 64)",
        "composición 65 a más": "composición de la población en años (porcentaje) (65 a más)",
        "fecundidad total": "tasa de fecundidad total",
        "poblacion": "población (miles)"
    }

    # Obtener todos los nombres de países en el archivo
    lista_paises = [row[0].lower() for row in sheet.iter_rows(min_row=3, max_row=77, min_col=1, max_col=1, values_only=True) if row[0]]
    
    # Identificar el país en la pregunta utilizando coincidencia difusa
    pais_mencionado, _ = process.extractOne(pregunta, lista_paises, scorer=fuzz.partial_ratio)

    # Identificar la categoría en la pregunta
    if "esperanza de vida" in pregunta:
        if "hombre" in pregunta:
            categoria = "esperanza de vida hombre"
        elif "mujer" in pregunta:
            categoria = "esperanza de vida mujer"
    elif "composición de la población" in pregunta:
        if "0 a 14" in pregunta:
            categoria = "composición 0 a 14"
        elif "15 a 64" in pregunta:
            categoria = "composición 15 a 64"
        elif "65 a más" in pregunta:
            categoria = "composición 65 a más"
    elif "fecundidad total" in pregunta or "tasa de fecundidad total" in pregunta:
        categoria = "fecundidad total"
    elif "poblacion" in pregunta:
        categoria = "poblacion"

    if not pais_mencionado or not categoria:
        return None  # Si no se encuentra país o categoría, devolver None

    # Obtener la respuesta específica según la categoría identificada
    for row in sheet.iter_rows(min_row=3, max_row=77, min_col=1, max_col=sheet.max_column, values_only=True):
        nombre_pais = row[0].lower()
        if pais_mencionado in nombre_pais:
            categoria_mapeada = categorias_mapeadas.get(categoria)
            if categoria_mapeada in encabezados:
                indice_categoria = encabezados.index(categoria_mapeada)

                # Obtener la respuesta específica para el país
                respuesta = row[indice_categoria]
                break

    return respuesta

if __name__ == "__main__":
    pregunta = sys.argv[1]
    respuesta = obtener_respuesta(pregunta)

    if respuesta is not None:
        print(respuesta)
    else:
        print("Lo siento, no encontre informacion relacionada con esa pregunta.")
